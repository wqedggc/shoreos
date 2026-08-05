#!/usr/bin/env python3
"""Local-only ledger import worker.

The worker is intentionally side-effect free: it reads one source document and
emits one JSON result to stdout.  The Go API owns files, authorization,
persistence, and user-scoped deduplication.
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import hashlib
import json
import re
import sys
from pathlib import Path
from typing import Any


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.date, dt.datetime)):
        if isinstance(value, dt.datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return value.isoformat()
    return str(value).strip().strip("\ufeff").rstrip("\t").strip()


def cents(value: str) -> int:
    normalized = clean(value).replace(",", "").replace("¥", "").replace("元", "")
    if not normalized:
        return 0
    negative = normalized.startswith("-")
    normalized = normalized.lstrip("+-")
    yuan, _, fraction = normalized.partition(".")
    result = int(yuan or "0") * 100 + int((fraction + "00")[:2])
    return -result if negative else result


def direction(raw: str) -> str:
    value = clean(raw)
    if "支出" in value or value in {"支", "付款"}:
        return "expense"
    if "收入" in value or value in {"收", "收款"}:
        return "income"
    return "neutral"


def signed_amount(direction_code: str, raw_amount: str) -> int:
    amount = abs(cents(raw_amount))
    return -amount if direction_code == "expense" else amount


def transaction_fingerprint(
    source: str,
    account_key: str,
    platform_order_no: str,
    merchant_order_no: str,
    raw_row_hash: str,
    occurrence: int,
) -> str:
    # A platform or merchant order number identifies the economic event even if
    # its status changes in a later export.  Without either ID, only an exactly
    # equal source row is idempotent; similar same-day transactions stay separate.
    if platform_order_no or merchant_order_no:
        values = [source, account_key, platform_order_no, merchant_order_no]
    else:
        values = [source, account_key, raw_row_hash]
    if occurrence > 1:
        values.append(str(occurrence))
    return sha256_bytes("\x1f".join(values).encode("utf-8"))


def detect_csv_encoding(path: Path) -> str:
    probe = path.read_bytes()[:8192]
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            probe.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            pass
    raise ValueError("无法识别 CSV 编码")


def parse_alipay(path: Path) -> dict[str, Any]:
    encoding = detect_csv_encoding(path)
    text = path.read_bytes().decode(encoding)
    lines = text.splitlines()
    try:
        header_index = next(index for index, line in enumerate(lines) if line.startswith("交易时间,"))
    except StopIteration as exc:
        raise ValueError("未找到支付宝交易表头") from exc

    def meta(pattern: str) -> str:
        matched = re.search(pattern, text)
        return matched.group(1).strip() if matched else ""

    account = meta(r"支付宝账户：(.+)")
    account_key = sha256_bytes(account.encode("utf-8")) if account else "unknown-alipay-account"
    reader = csv.DictReader(lines[header_index:])
    transactions = []
    occurrences: dict[str, int] = {}
    for raw in reader:
        row = {clean(key): clean(value) for key, value in raw.items() if key}
        transaction_time = row.get("交易时间", "")
        if not re.match(r"20\d{2}-\d{2}-\d{2}", transaction_time):
            continue
        direction_code = direction(row.get("收/支", ""))
        amount_cents = signed_amount(direction_code, row.get("金额", ""))
        platform_order_no = row.get("交易订单号", "")
        merchant_order_no = row.get("商家订单号", "")
        counterparty = row.get("交易对方", "")
        product = row.get("商品说明", "")
        status = row.get("交易状态", "")
        raw_line = ",".join(row.get(clean(field), "") for field in (reader.fieldnames or []))
        raw_row_hash = sha256_bytes(raw_line.encode("utf-8"))
        occurrences[raw_row_hash] = occurrences.get(raw_row_hash, 0) + 1
        occurrence = occurrences[raw_row_hash]
        transactions.append({
            "occurredAt": transaction_time,
            "category": row.get("交易分类", ""),
            "counterpartyName": counterparty,
            "counterpartyAccount": row.get("对方账号", ""),
            "productDescription": product,
            "direction": direction_code,
            "amountCents": amount_cents,
            "paymentMethod": row.get("收/付款方式", ""),
            "status": status,
            "platformOrderNo": platform_order_no,
            "merchantOrderNo": merchant_order_no,
            "remark": row.get("备注", ""),
            "rawSnapshot": row,
            "rawRowHash": raw_row_hash,
            "transactionFingerprint": transaction_fingerprint(
                "alipay", account_key, platform_order_no, merchant_order_no, raw_row_hash, occurrence,
            ),
        })
    return {
        "sourceType": "alipay_csv",
        "parserVersion": "2",
        "statement": {
            "platform": "alipay",
            "accountKey": account_key,
            "startAt": meta(r"起始时间：\[(.+?)\]"),
            "endAt": meta(r"终止时间：\[(.+?)\]"),
            "exportedAt": meta(r"导出时间：\[(.+?)\]"),
        },
        "transactions": transactions,
    }


def locate_wechat_header(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    required = {"交易时间", "收/支"}
    for index, row in enumerate(rows[:40]):
        headers = [clean(value) for value in row]
        present = set(headers)
        if required.issubset(present) and ("金额(元)" in present or "金额" in present):
            return index, headers
    raise ValueError("未找到微信交易表头")


def parse_wechat(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("缺少 openpyxl，无法解析微信 Excel") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    all_rows = list(worksheet.iter_rows(values_only=True))
    header_index, headers = locate_wechat_header(all_rows)
    meta_text = "\n".join(clean(row[0]) for row in all_rows[:header_index] if row)

    def meta(pattern: str) -> str:
        matched = re.search(pattern, meta_text)
        return matched.group(1).strip() if matched else ""

    nickname = meta(r"微信昵称：\[(.*?)\]")
    account_key = sha256_bytes((nickname or "unknown-wechat-account").encode("utf-8"))
    transactions = []
    occurrences: dict[str, int] = {}
    for values in all_rows[header_index + 1:]:
        row = {headers[index]: clean(value) for index, value in enumerate(values) if index < len(headers) and headers[index]}
        transaction_time = row.get("交易时间", "")
        if not re.match(r"20\d{2}-\d{2}-\d{2}", transaction_time):
            continue
        direction_code = direction(row.get("收/支", ""))
        amount_cents = signed_amount(direction_code, row.get("金额(元)", row.get("金额", "")))
        platform_order_no = row.get("交易单号", "")
        merchant_order_no = row.get("商户单号", "")
        counterparty = row.get("交易对方", "")
        product = row.get("商品", "")
        status = row.get("当前状态", "")
        raw_line = ",".join(row.get(header, "") for header in headers)
        raw_row_hash = sha256_bytes(raw_line.encode("utf-8"))
        occurrences[raw_row_hash] = occurrences.get(raw_row_hash, 0) + 1
        occurrence = occurrences[raw_row_hash]
        transactions.append({
            "occurredAt": transaction_time,
            "category": row.get("交易类型", ""),
            "counterpartyName": counterparty,
            "counterpartyAccount": "",
            "productDescription": product,
            "direction": direction_code,
            "amountCents": amount_cents,
            "paymentMethod": row.get("支付方式", ""),
            "status": status,
            "platformOrderNo": platform_order_no,
            "merchantOrderNo": merchant_order_no,
            "remark": row.get("备注", ""),
            "rawSnapshot": row,
            "rawRowHash": raw_row_hash,
            "transactionFingerprint": transaction_fingerprint(
                "wechat", account_key, platform_order_no, merchant_order_no, raw_row_hash, occurrence,
            ),
        })
    range_match = re.search(r"起始时间：\[(.*?)\]\s*终止时间：\[(.*?)\]", meta_text)
    return {
        "sourceType": "wechat_xlsx",
        "parserVersion": "2",
        "statement": {
            "platform": "wechat",
            "accountKey": account_key,
            "startAt": range_match.group(1).strip() if range_match else "",
            "endAt": range_match.group(2).strip() if range_match else "",
            "exportedAt": meta(r"导出时间：\[(.*?)\]"),
        },
        "transactions": transactions,
    }


CMB_ROW = re.compile(
    r"^(?P<date>20\d{2}-\d{2}-\d{2})\s+(?P<currency>[A-Z]{3})\s+"
    r"(?P<amount>-?\d{1,3}(?:,\d{3})*(?:\.\d{2})|-?\d+\.\d{2})\s+"
    r"(?P<balance>-?\d{1,3}(?:,\d{3})*(?:\.\d{2})|-?\d+\.\d{2})\s+(?P<rest>.+)$"
)


def parse_cmb_pdf(path: Path) -> dict[str, Any]:
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValueError("缺少 pypdf，无法解析招商 PDF") from exc

    text = "\n".join(page.extract_text() or "" for page in PdfReader(path).pages)
    date_range = re.search(r"(20\d{2}-\d{2}-\d{2})\s*--\s*(20\d{2}-\d{2}-\d{2})", text)
    account_match = re.search(r"账号：\s*([0-9]+)", text)
    if not date_range or not account_match:
        raise ValueError("招商 PDF 缺少日期范围或账号")
    account_key = sha256_bytes(account_match.group(1).encode("utf-8"))
    rows = []
    current: dict[str, Any] | None = None
    occurrences: dict[str, int] = {}

    def flush() -> None:
        nonlocal current
        if not current:
            return
        rest = current["rest"].strip()
        parts = rest.split(None, 1)
        transaction_type = parts[0] if parts else ""
        counterparty = parts[1] if len(parts) > 1 else ""
        amount_cents = cents(current["amount"])
        occurred_at = current["date"]
        raw_line = " ".join(current["lines"])
        raw_row_hash = sha256_bytes(raw_line.encode("utf-8"))
        occurrences[raw_row_hash] = occurrences.get(raw_row_hash, 0) + 1
        occurrence = occurrences[raw_row_hash]
        rows.append({
            "occurredAt": occurred_at,
            "category": transaction_type,
            "counterpartyName": counterparty,
            "counterpartyAccount": "",
            "productDescription": "",
            "direction": "income" if amount_cents >= 0 else "expense",
            "amountCents": amount_cents,
            "paymentMethod": "",
            "status": "completed",
            "platformOrderNo": "",
            "merchantOrderNo": "",
            "remark": "",
            "rawSnapshot": {"currency": current["currency"], "balanceCents": cents(current["balance"]), "transactionType": transaction_type},
            "rawRowHash": raw_row_hash,
            "transactionFingerprint": transaction_fingerprint(
                "cmb_bank_pdf", account_key, "", "", raw_row_hash, occurrence,
            ),
        })
        current = None

    for raw_line in text.splitlines():
        line = re.sub(r"\s+", " ", raw_line.strip())
        if not line or re.match(r"^\d+/\d+$", line) or line.startswith(("招商银行交易流水", "记账日期", "Date Currency", "Amount Balance", "申请时间", "账号", "户名")):
            continue
        match = CMB_ROW.match(line)
        if match:
            flush()
            current = {**match.groupdict(), "lines": [line]}
        elif current:
            current["rest"] = f"{current['rest']} {line}".strip()
            current["lines"].append(line)
    flush()
    return {
        "sourceType": "cmb_bank_pdf_text",
        "parserVersion": "2",
        "statement": {"platform": "cmb_bank", "accountKey": account_key, "startAt": date_range.group(1), "endAt": date_range.group(2), "exportedAt": ""},
        "transactions": rows,
    }


def detect_source_type(path: Path) -> str:
    suffix = path.suffix.lower()
    source_type = {".csv": "alipay_csv", ".xlsx": "wechat_xlsx", ".pdf": "cmb_bank_pdf_text"}.get(suffix, "")
    if source_type:
        return source_type
    signature = path.read_bytes()[:8]
    if signature.startswith(b"%PDF-"):
        return "cmb_bank_pdf_text"
    if signature.startswith(b"PK\x03\x04"):
        return "wechat_xlsx"
    return "alipay_csv"


def parse(path: Path, source_type: str) -> dict[str, Any]:
    if source_type == "auto":
        source_type = detect_source_type(path)
    parsers = {"alipay_csv": parse_alipay, "wechat_xlsx": parse_wechat, "cmb_bank_pdf_text": parse_cmb_pdf}
    parser = parsers.get(source_type)
    if not parser:
        raise ValueError("不支持的账单格式")
    result = parser(path)
    result["file"] = {"name": path.name, "sha256": sha256_file(path), "bytes": path.stat().st_size}
    result["transactionCount"] = len(result["transactions"])
    return result


def main() -> int:
    arguments = argparse.ArgumentParser()
    arguments.add_argument("file")
    arguments.add_argument("--source-type", default="auto", choices=("auto", "alipay_csv", "wechat_xlsx", "cmb_bank_pdf_text"))
    args = arguments.parse_args()
    try:
        result = parse(Path(args.file).expanduser().resolve(), args.source_type)
        print(json.dumps(result, ensure_ascii=False, separators=(",", ":")))
        return 0
    except Exception as exc:  # stdout remains machine-readable and non-sensitive.
        print(json.dumps({"error": {"code": "PARSE_FAILED", "message": str(exc)}}, ensure_ascii=False))
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
